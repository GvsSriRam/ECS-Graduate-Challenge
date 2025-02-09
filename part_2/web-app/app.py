import os
import sqlite3
import json
import random
import smtplib
from email.mime.text import MIMEText
from io import StringIO, BytesIO
import csv
import openpyxl
from flask import Flask, g, request, redirect, url_for, session, flash, send_file, render_template_string
from flask_mail import Mail, Message
import qrcode

# ===============================
# Configuration and Initialization
# ===============================
app = Flask(__name__)
app.secret_key = 'your-secret-key-here'  # Replace with a secure random key in production
DATABASE = 'app.db'

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DATABASE)
        db.row_factory = sqlite3.Row
    return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

# def init_db():
#     """Initialize the database with judges, scores, and score_changes tables."""
#     with app.app_context():
#         db = get_db()
#         db.execute('''
#             CREATE TABLE IF NOT EXISTS judges (
#                 email TEXT PRIMARY KEY,
#                 name TEXT NOT NULL,
#                 pin TEXT,
#                 assigned_posters TEXT,  -- Stored as JSON list of poster IDs
#                 assigned_poster_titles TEXT  -- Stored as JSON object mapping poster IDs to poster titles
#             )
#         ''')
#         db.execute('''
#             CREATE TABLE IF NOT EXISTS scores (
#                 id INTEGER PRIMARY KEY AUTOINCREMENT,
#                 judge_email TEXT,
#                 poster_id TEXT,
#                 score INTEGER,
#                 UNIQUE(judge_email, poster_id)
#             )
#         ''')
#         db.execute('''
#             CREATE TABLE IF NOT EXISTS score_changes (
#                 id INTEGER PRIMARY KEY AUTOINCREMENT,
#                 judge_email TEXT,
#                 poster_id TEXT,
#                 old_score INTEGER,
#                 new_score INTEGER,
#                 change_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP
#             )
#         ''')
#         db.commit()

def init_db():
    """Initialize the database with judges, judge_info, scores, and score_changes tables."""
    with app.app_context():
        db = get_db()
        # Judges table stores poster assignments
        db.execute('''
            CREATE TABLE IF NOT EXISTS judges (
                email TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                pin TEXT,
                assigned_posters TEXT,  -- Stored as JSON list of poster IDs; can be empty "[]"
                assigned_poster_titles TEXT  -- Stored as JSON mapping poster IDs to titles
            )
        ''')
        # judge_info table: judge_id is provided (from the import file)
        db.execute('''
            CREATE TABLE IF NOT EXISTS judge_info (
                judge_id INTEGER PRIMARY KEY,
                name TEXT NOT NULL,
                email TEXT UNIQUE NOT NULL
            )
        ''')
        db.execute('''
            CREATE TABLE IF NOT EXISTS scores (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                judge_email TEXT,
                poster_id TEXT,
                score INTEGER,
                UNIQUE(judge_email, poster_id)
            )
        ''')
        db.execute('''
            CREATE TABLE IF NOT EXISTS score_changes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                judge_email TEXT,
                poster_id TEXT,
                old_score INTEGER,
                new_score INTEGER,
                change_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        db.commit()



# Initialize DB on first run
if not os.path.exists(DATABASE):
    init_db()
    # Insert sample judge data for demo purposes
    db = sqlite3.connect(DATABASE)
    sample_judges = [
        ("judge1@example.com", "Judge One", json.dumps(["poster1", "poster2", "poster3"])),
        ("judge2@example.com", "Judge Two", json.dumps(["poster2", "poster4"]))
    ]
    db.executemany("INSERT OR IGNORE INTO judges (email, name, assigned_posters) VALUES (?, ?, ?)", sample_judges)
    db.commit()
    db.close()

# In-memory OTP store: mapping email -> OTP (for a 24-hour hackathon prototype, this is acceptable)
otp_store = {}

# ===============================
# Flask-Mail Configuration
# ===============================
app.config.update(
    MAIL_SERVER='smtp.gmail.com',
    MAIL_PORT=587,
    MAIL_USE_TLS=True,
    MAIL_USERNAME='rajmohan8081@gmail.com',        # Replace with your email address
    MAIL_PASSWORD='ygxc olrd xjhm gawf',             # Replace with your app-specific password
    MAIL_DEFAULT_SENDER='rajmohan8081@gmail.com'
)

mail = Mail(app)

def send_otp(email, otp):
    """
    Sends an OTP email using Flask-Mail.
    """
    msg = Message("Your Login OTP", recipients=[email])
    msg.body = f"Your OTP for login is: {otp}"
    try:
        mail.send(msg)
        print(f"Sent OTP to {email}")
    except Exception as e:
        print(f"Error sending email: {e}")

# ===============================
# Root Route for Easy Navigation
# ===============================
@app.route('/')
def index():
    """Redirect to dashboard if logged in, else to login page."""
    if 'user' in session:
        return redirect(url_for('dashboard'))
    else:
        return redirect(url_for('login'))

# ===============================
# Routes for Authentication
# ===============================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """
    Dark-themed login screen for judges.
    A judge is allowed to login only if they have at least one poster assignment.
    """
    if request.method == 'POST':
        email = request.form.get('email')
        db = get_db()
        cur = db.execute("SELECT * FROM judges WHERE email = ?", (email,))
        judge = cur.fetchone()
        if judge:
            # Check if the judge has any assigned posters
            assigned_posters = json.loads(judge['assigned_posters']) if judge['assigned_posters'] else []
            if not assigned_posters:
                flash("You do not have any poster assignments and cannot log in.")
                return redirect(url_for('login'))
            # Otherwise, generate a 6-digit OTP and send it
            otp = str(random.randint(100000, 999999))
            otp_store[email] = otp
            send_otp(email, otp)
            session['pending_email'] = email
            flash("An OTP has been sent to your email. Please enter it below.")
            return redirect(url_for('verify'))
        else:
            flash("Email not authorized.")
            return redirect(url_for('login'))

    # Render a dark-themed login page
    return render_template_string('''
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"/>
  <title>Login</title>
  <meta name="viewport" content="width=device-width, initial-scale=1.0"/>
  <!-- IBM Plex Sans via Google Fonts -->
  <link rel="stylesheet"
        href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;600&display=swap"/>
  <style>
    /* Your existing CSS styles */
    * { box-sizing: border-box; margin: 0; padding: 0; font-family: 'IBM Plex Sans', sans-serif; }
    body { background-color: #121212; color: #fff; display: flex; flex-direction: column; align-items: center; justify-content: center; min-height: 100vh; }
    .card { background-color: #1f1f1f; border-radius: 20px; width: 90%; max-width: 400px; padding: 2rem; }
    .title { font-size: 1.5rem; font-weight: 600; margin-bottom: 1.5rem; text-align: center; }
    .input-group { margin-bottom: 1rem; }
    label { display: block; font-size: 0.9rem; margin-bottom: 0.4rem; color: #ccc; }
    input[type="email"] { width: 100%; padding: 0.8rem; border-radius: 10px; border: none; outline: none; font-size: 1rem; background-color: #2a2a2a; color: #fff; }
    .btn { display: block; width: 100%; background-color: #4285F4; color: #fff; font-size: 1rem; font-weight: 600; text-align: center; padding: 0.8rem; border: none; border-radius: 10px; margin-top: 1rem; cursor: pointer; }
    .btn:hover { background-color: #357ae8; }
    .info { font-size: 0.85rem; color: #aaa; margin-top: 0.5rem; text-align: center; }
  </style>
</head>
<body>
  <div class="card">
    <div class="title">ECS Research Day Portal</div>
    <form method="post">
      <div class="input-group">
        <label for="email">Enter your email</label>
        <input type="email" id="email" name="email" placeholder="your.email@syr.edu" required />
      </div>
      <button type="submit" class="btn">Send OTP</button>
      <div class="info">Build for Syracuse University 🍊</div>
    </form>
  </div>
</body>
</html>
    ''')


@app.route('/verify', methods=['GET', 'POST'])
def verify():
    pending_email = session.get('pending_email')
    if not pending_email:
        return redirect(url_for('login'))

    if request.method == 'POST':
        otp_input = (
            request.form.get('otp1', '') +
            request.form.get('otp2', '') +
            request.form.get('otp3', '') +
            request.form.get('otp4', '') +
            request.form.get('otp5', '') +
            request.form.get('otp6', '')
        )
        
        if otp_store.get(pending_email) == otp_input:
            session['user'] = pending_email
            otp_store.pop(pending_email, None)
            flash("Logged in successfully.")
            # Check if there is a redirect URL saved from a QR scan
            redirect_url = session.pop('redirect_after_login', None)
            if redirect_url:
                return redirect(redirect_url)
            else:
                return redirect(url_for('dashboard'))
        else:
            flash("Invalid OTP. Please try again.")
            return redirect(url_for('verify'))

    # Render the 6-digit OTP form
    return render_template_string('''
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"/>
  <title>Verify OTP</title>
  <meta name="viewport" content="width=device-width, initial-scale=1.0"/>
  <!-- IBM Plex Sans via Google Fonts (or any other) -->
  <link rel="stylesheet"
        href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;600&display=swap"/>
  <style>
    * {
      box-sizing: border-box;
      margin: 0;
      padding: 0;
      font-family: 'IBM Plex Sans', sans-serif;
    }
    body {
      background-color: #121212;
      color: #fff;
      display: flex;
      align-items: center;
      justify-content: center;
      height: 100vh;
    }
    .card {
      background-color: #1f1f1f;
      padding: 2rem;
      border-radius: 16px;
      width: 90%;
      max-width: 400px;
      text-align: center;
    }
    .card h2 {
      font-size: 1.5rem;
      margin-bottom: 0.5rem;
    }
    .card p {
      color: #bbb;
      margin-bottom: 1.5rem;
    }
    form {
      display: flex;
      flex-direction: column;
      align-items: center;
    }
    .otp-container {
      display: flex;
      gap: 0.5rem;
      margin-bottom: 1rem;
    }
    .otp-input {
      width: 3rem;
      height: 3rem;
      font-size: 1.5rem;
      text-align: center;
      border-radius: 8px;
      border: none;
      background-color: #2a2a2a;
      color: #fff;
    }
    .submit-btn {
      background-color: #4285F4;
      color: #fff;
      font-size: 1rem;
      padding: 0.8rem 1.5rem;
      border: none;
      border-radius: 8px;
      cursor: pointer;
    }
    .submit-btn:hover {
      background-color: #357ae8;
    }
    .resend-link {
      margin-top: 1rem;
      color: #4285F4;
      text-decoration: underline;
      cursor: pointer;
      font-size: 0.9rem;
    }
  </style>
</head>
<body>
  <div class="card">
    <h2>Verify OTP</h2>
    <p>Code has been sent to {{ pending_email }}</p>
    <form method="post" id="otpForm">
      <div class="otp-container">
        <!-- 6 separate inputs, each 1 digit -->
        {% for i in range(1,7) %}
        <input
          type="tel"
          inputmode="numeric"
          maxlength="1"
          class="otp-input"
          name="otp{{ i }}"
          id="otp{{ i }}"
          required
        />
        {% endfor %}
      </div>
      <button type="submit" class="submit-btn">Verify</button>
    </form>
    <div 
      class="resend-link" 
      onclick="window.location.href='/login';">
      Didn’t get OTP? Resend Code
    </div>
  </div>

  <script>
    // Focus the first input on load
    document.getElementById('otp1').focus();

    // For all 6 inputs, jump to next on keyup if a single digit is entered
    const inputs = document.querySelectorAll('.otp-input');
    inputs.forEach((input, index) => {
      input.addEventListener('input', (e) => {
        if (input.value.length === 1) {
          // Move to next input if it exists
          if (index < inputs.length - 1) {
            inputs[index + 1].focus();
          }
        }
      });

      // Handle backspace to move focus back
      input.addEventListener('keydown', (e) => {
        if (e.key === 'Backspace' && input.value === '' && index > 0) {
          inputs[index - 1].focus();
        }
      });
    });
  </script>
</body>
</html>
    ''', pending_email=pending_email)



@app.route('/logout')
def logout():
    session.clear()
    flash("Logged out.")
    return redirect(url_for('login'))


@app.route('/dashboard')
def dashboard():
    """
    Dark-theme dashboard that greets the judge and shows the assigned posters.
    Displays each poster in the format "poster_number: Poster_name".
    """
    if 'user' not in session:
        return redirect(url_for('login'))
    
    email = session['user']
    db = get_db()
    cur = db.execute("SELECT * FROM judges WHERE email = ?", (email,))
    judge = cur.fetchone()
    if not judge:
        flash("Judge not found.")
        return redirect(url_for('login'))
    
    # Load the assigned posters (list of poster ids) and the poster title mapping.
    assigned_posters = json.loads(judge['assigned_posters']) if judge['assigned_posters'] else []
    poster_titles = json.loads(judge['assigned_poster_titles']) if judge['assigned_poster_titles'] else {}

    # Retrieve existing scores for these posters.
    scores = {}
    for poster in assigned_posters:
        cur = db.execute(
            "SELECT score FROM scores WHERE judge_email = ? AND poster_id = ?",
            (email, poster)
        )
        row = cur.fetchone()
        scores[poster] = row['score'] if row else None

    # Render the dashboard template.
    dashboard_html = render_template_string('''
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0"/>
  <title>Dashboard</title>
  <!-- Include IBM Plex Sans via Google Fonts -->
  <link rel="stylesheet"
        href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;600&display=swap"/>
  <style>
    * { box-sizing: border-box; margin: 0; padding: 0; font-family: 'IBM Plex Sans', sans-serif; }
    body { background-color: #121212; color: #fff; }
    .header { background-color: #1f1f1f; border-bottom-left-radius: 24px; border-bottom-right-radius: 24px; padding: 2rem 1.5rem; }
    .header .top-row { display: flex; align-items: center; justify-content: space-between; }
    .header .greeting h2 { font-size: 1.2rem; font-weight: 600; margin-bottom: 0.2rem; }
    .header .greeting span { font-size: 0.9rem; color: #ccc; }
    .main-content { margin-top: 1rem; padding: 1.5rem; }
    .poster-list { display: flex; flex-direction: column; gap: 1rem; }
    .poster-card { background-color: #1f1f1f; border-radius: 12px; padding: 1rem; display: flex; align-items: center; justify-content: space-between; }
    .poster-info { display: flex; flex-direction: column; }
    .poster-name { font-weight: 600; margin-bottom: 0.3rem; }
    .poster-score { color: #999; font-size: 0.9rem; }
    .score-link { color: #fff; background-color: #4285F4; padding: 0.4rem 0.6rem; border-radius: 8px; text-decoration: none; font-size: 0.85rem; }
    .score-link:hover { background-color: #357ae8; }
    .bottom-links { display: flex; justify-content: space-around; margin-top: 2rem; }
    .bottom-links a { text-decoration: none; color: #fff; background: #4285F4; padding: 0.6rem 1.2rem; border-radius: 8px; transition: background 0.2s ease-in-out; }
    .bottom-links a:hover { background: #357ae8; }
    .top-row a {
      text-decoration: none;
      color: #fff;
      background: red;
      padding: 0.6rem 1.2rem;
      border-radius: 8px;
      transition: background 0.2s ease-in-out;
    }
  </style>
</head>
<body>
  <!-- Header -->
  <div class="header">
    <div class="top-row">
      <div class="greeting">
        <h2>Hi &#128075; {{ judge.name }}!</h2>
        <span>Welcome back</span>
      </div>
      <div class="top-row">
        <a href="{{ url_for('logout') }}">Logout</a>
      </div>
    </div>
  </div>

  <!-- Main content -->
  <div class="main-content">
    <div class="poster-list">
      {% for poster in assigned_posters %}
      <div class="poster-card">
        <div class="poster-info">
          <!-- Extract poster number from the poster id and show in the format: poster_number: Poster_name -->
          <span class="poster-name">
            Poster-{{ poster[6:] }}: {{ poster_titles.get(poster, 'No Title') }}
          </span>
          {% if scores[poster] is not none %}
            <span class="poster-score">Score: {{ scores[poster] }}</span>
          {% else %}
            <span class="poster-score">Not Scored</span>
          {% endif %}
        </div>
        <a class="score-link" href="{{ url_for('score', poster_id=poster) }}">Update</a>
      </div>
      {% endfor %}
    </div>

    <div class="bottom-links">
      <a href="{{ url_for('score_log') }}">Score Log</a>
    </div>
  </div>
</body>
</html>
    ''',
    judge=judge,
    assigned_posters=assigned_posters,
    poster_titles=poster_titles,
    scores=scores
    )
    
    return dashboard_html


@app.route('/score/<poster_id>', methods=['GET', 'POST'])
def score(poster_id):
    """
    Allows a judge to enter or update a score for an assigned poster.
    Displays the poster in the format "poster_number: Poster_name" at the top.
    Uses a mobile-optimized UI and follows the dark theme.
    """
    if 'user' not in session:
        return redirect(url_for('login'))
    email = session['user']
    db = get_db()
    # Fetch both assigned_posters and assigned_poster_titles
    cur = db.execute("SELECT assigned_posters, assigned_poster_titles FROM judges WHERE email = ?", (email,))
    row = cur.fetchone()
    if not row:
        flash("Judge not found.")
        return redirect(url_for('login'))
    assigned_posters = json.loads(row['assigned_posters']) if row['assigned_posters'] else []
    poster_titles = json.loads(row['assigned_poster_titles']) if row['assigned_poster_titles'] else {}
    if poster_id not in assigned_posters:
        flash("You are not assigned to this poster.")
        return redirect(url_for('dashboard'))
    
    # Compute display string in the format "poster_number: Poster_name"
    poster_display = f"{poster_id[6:]}: {poster_titles.get(poster_id, 'No Title')}"
    
    # Get the current score if it exists
    cur = db.execute("SELECT score FROM scores WHERE judge_email = ? AND poster_id = ?", (email, poster_id))
    row = cur.fetchone()
    current_score = row['score'] if row else ""
    
    if request.method == 'POST':
        score_val = request.form.get('score')
        try:
            score_int = int(score_val)
            if score_int < 0 or score_int > 10:
                flash("Score must be between 0 and 10.")
                return redirect(url_for('score', poster_id=poster_id))
        except ValueError:
            flash("Invalid score.")
            return redirect(url_for('score', poster_id=poster_id))
        
        # Check if there is an existing score
        cur = db.execute("SELECT score FROM scores WHERE judge_email = ? AND poster_id = ?", (email, poster_id))
        row = cur.fetchone()
        old_score = row['score'] if row else None

        # Insert or update the score
        db.execute('''INSERT INTO scores (judge_email, poster_id, score)
                      VALUES (?, ?, ?)
                      ON CONFLICT(judge_email, poster_id) DO UPDATE SET score = ?''',
                   (email, poster_id, score_int, score_int))
        db.commit()
        
        # Log the change if needed
        if old_score is None or old_score != score_int:
            db.execute('''INSERT INTO score_changes (judge_email, poster_id, old_score, new_score)
                          VALUES (?, ?, ?, ?)''', (email, poster_id, old_score, score_int))
            db.commit()
        
        flash("Score updated.")
        return redirect(url_for('dashboard'))
    
    return render_template_string('''
<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Enter Score for {{ poster_display }}</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
  <style>
    body { background-color: #121212; font-family: 'Roboto', sans-serif; color: white; }
    .card { border-radius: 20px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); background: #1f1f1f; }
    .score-input { font-size: 2rem; text-align: center; border: none; outline: none; width: 100%; background-color: transparent; color: white; }
    .submit-btn { background-color: #4285F4; border: none; color: white; font-size: 1.5rem; border-radius: 10px; padding: 10px 0; margin-top: 20px; width: 100%; }
    .submit-btn:hover { background-color: #357ae8; }
    .text-color-w { color: white; }
    .back-btn { padding: .5em; background-color: #4285F4; text-decoration: none; color: white; border-radius: 10px; }
  </style>
</head>
<body>
  <div class="container d-flex justify-content-center align-items-center vh-100">
    <div class="card p-4" style="width: 90%; max-width: 400px;">
      <h2 class="text-color-w text-center mb-4">Enter Score</h2>
      <!-- Display the poster number and title in the desired format -->
      <h5 class="text-center mb-4 text-color-w">{{ poster_display }}</h5>
      <form method="post">
        <div class="mb-3">
          <input type="number" name="score" min="0" max="10" value="{{ current_score }}" class="score-input" placeholder="0-10" required>
        </div>
        <button type="submit" class="submit-btn">Submit</button>
      </form>
      <a href="{{ url_for('dashboard') }}" class="btn btn-link d-block text-center mt-3 back-btn">Back</a>
    </div>
  </div>
</body>
</html>
''', poster_display=poster_display, current_score=current_score)



# ===============================
# Score Change Log for Judges
# ===============================

@app.route('/score_log')
def score_log():
    """
    Display the log of score changes as a dark-themed timeline,
    matching the style of the main dashboard screen.
    """
    if 'user' not in session:
        return redirect(url_for('login'))
    
    email = session['user']
    db = get_db()
    # Fetch changes in descending chronological order
    cur = db.execute("""
        SELECT * FROM score_changes 
        WHERE judge_email = ? 
        ORDER BY change_time DESC
    """, (email,))
    changes = cur.fetchall()

    # Render a timeline UI with dark theme and IBM Plex Sans
    log_html = render_template_string('''
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <title>Score Log</title>
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <!-- IBM Plex Sans -->
  <link rel="stylesheet"
        href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;600&display=swap">
  <style>
    * {
      box-sizing: border-box;
      margin: 0;
      padding: 0;
      font-family: 'IBM Plex Sans', sans-serif;
    }
    body {
      background-color: #121212;
      color: #fff;
      padding: 1rem;
    }
    .header {
      background-color: #1f1f1f;
      border-radius: 16px;
      padding: 1.5rem;
      margin-bottom: 1.5rem;
    }
    .header h2 {
      font-weight: 600;
      font-size: 1.2rem;
    }
    .timeline {
      position: relative;
      margin-left: 2rem; /* space for the vertical line */
      padding: 0.5rem 0;
      border-left: 2px solid #2a2a2a; /* the timeline line */
    }
    .timeline-item {
      margin-bottom: 1.5rem;
      position: relative;
    }
    .timeline-item:last-child {
      margin-bottom: 0;
    }
    .timeline-dot {
      position: absolute;
      left: -0.53rem; /* horizontally center the dot on the line */
      top: 0.45rem;
      width: 12px;
      height: 12px;
      border-radius: 50%;
      background-color: #4285F4;
      border: 2px solid #121212; /* so it stands out from the line if needed */
    }
    .item-content {
      padding-left: 1.5rem;
    }
    .time-stamp {
      font-size: 0.85rem;
      color: #999;
      margin-bottom: 0.2rem;
    }
    .poster-id {
      font-weight: 600;
      color: #fff;
      margin-bottom: 0.2rem;
    }
    .score-change {
      font-size: 0.9rem;
      color: #ccc;
    }
    .score-change span {
      color: #76c893; /* greenish or any highlight color for new score */
    }
    .no-changes {
      font-style: italic;
      color: #ccc;
      margin-top: 2rem;
      text-align: center;
    }
    .nav-links {
      text-align: center;
      margin-top: 2rem;
    }
    .nav-links a {
      text-decoration: none;
      color: #fff;
      background: #4285F4;
      padding: 0.6rem 1.2rem;
      border-radius: 8px;
      transition: background 0.2s ease-in-out;
      margin: 0 0.5rem;
    }
    .nav-links a:hover {
      background: #357ae8;
    }
  </style>
</head>
<body>
  <div class="header">
    <h2>Your Score Change Log</h2>
  </div>

  {% if not changes %}
    <p class="no-changes">No score changes logged yet.</p>
  {% else %}
    <div class="timeline">
      {% for change in changes %}
        <div class="timeline-item">
          <div class="timeline-dot"></div>
          <div class="item-content">
            <div class="time-stamp">{{ change['change_time'] }}</div>
            <div class="poster-id">Poster: {{ change['poster_id'] }}</div>
            <div class="score-change">
              <!-- If old_score is None, it means first-time score creation -->
              {% if change['old_score'] is none %}
                Initial score set to <span>{{ change['new_score'] }}</span>
              {% else %}
                Changed from <span>{{ change['old_score'] }}</span> to <span>{{ change['new_score'] }}</span>
              {% endif %}
            </div>
          </div>
        </div>
      {% endfor %}
    </div>
  {% endif %}

  <div class="nav-links">
    <a href="{{ url_for('dashboard') }}">Back to Dashboard</a>
    
  </div>
</body>
</html>
    ''', changes=changes)

    return log_html

@app.route('/admin/remove_judges', methods=['GET', 'POST'])
def remove_judges():
    """
    Admin-only endpoint (accessible only if ?key=adminsecret is provided)
    to remove all current judges from the database.
    
    GET: Displays a dark-themed confirmation form.
    POST: Deletes all rows from the judges table.
    """
    if request.args.get('key') != 'adminsecret':
        return "Unauthorized", 401

    if request.method == 'POST':
        db = get_db()
        db.execute("DELETE FROM judges")
        db.commit()
        return "All judges have been removed.", 200

    return render_template_string('''
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Remove Judges</title>
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <link rel="stylesheet"
                  href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;600&display=swap">
            <style>
                body {
                  background-color: #121212;
                  color: #fff;
                  font-family: 'IBM Plex Sans', sans-serif;
                  margin: 0;
                  padding: 2rem;
                }
                .container {
                  max-width: 500px;
                  margin: auto;
                  background: #1f1f1f;
                  padding: 2rem;
                  border-radius: 12px;
                  box-shadow: 0 4px 6px rgba(0,0,0,0.3);
                  text-align: center;
                }
                h2 { margin-bottom: 1rem; }
                p { margin-bottom: 2rem; }
                button {
                  padding: 0.5rem 1rem;
                  background: #d9534f;
                  color: #fff;
                  border: none;
                  border-radius: 8px;
                  cursor: pointer;
                  font-size: 1rem;
                }
                button:hover {
                  background: #c9302c;
                }
                a.btn {
                  display: inline-block;
                  margin-top: 1rem;
                  padding: 0.5rem 1rem;
                  background: #4285F4;
                  color: #fff;
                  border-radius: 8px;
                  text-decoration: none;
                  font-size: 1rem;
                }
                a.btn:hover {
                  background: #357ae8;
                }
            </style>
        </head>
        <body>
            <div class="container">
                <h2>Remove All Judges</h2>
                <p>Are you sure you want to remove all judges from the system?<br>
                   This action cannot be undone.</p>
                <form method="POST">
                    <button type="submit">Remove Judges</button>
                </form>
                <a href="{{ url_for('admin_dashboard') }}?key=adminsecret" class="btn">Cancel</a>
            </div>
        </body>
        </html>
    ''')


# def is_valid_poster(cell_value):
#     """Return True if the cell value represents a valid poster number."""
#     try:
#         return cell_value is not None and (isinstance(cell_value, (int, float)) or str(cell_value).strip() != "")
#     except Exception:
#         return False

# @app.route('/admin/import_judges', methods=['GET', 'POST'])
# def import_judges():
#     """
#     Admin-only endpoint (accessible only if ?key=adminsecret is provided)
#     that accepts an XLSX file upload containing judge and poster assignments.
    
#     The XLSX file is expected to have a header row with columns like:
#       - Email, Judge FirstName, Judge LastName, poster-1, poster-1-title, ... poster-6, poster-6-title
      
#     Judges with no valid poster assignments are skipped.
#     """
#     if request.args.get('key') != 'adminsecret':
#         return "Unauthorized", 401

#     if request.method == 'GET':
#         return render_template_string('''
#         <!DOCTYPE html>
#         <html>
#         <head>
#             <meta charset="UTF-8">
#             <title>Import Judges</title>
#             <meta name="viewport" content="width=device-width, initial-scale=1.0">
#             <link rel="stylesheet"
#                   href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;600&display=swap">
#             <style>
#                 body {
#                   background-color: #121212;
#                   color: #fff;
#                   font-family: 'IBM Plex Sans', sans-serif;
#                   margin: 0;
#                   padding: 2rem;
#                 }
#                 .container {
#                   max-width: 500px;
#                   margin: auto;
#                   background: #1f1f1f;
#                   padding: 2rem;
#                   border-radius: 12px;
#                   box-shadow: 0 4px 6px rgba(0,0,0,0.3);
#                   text-align: center;
#                 }
#                 input[type="file"] {
#                   display: block;
#                   margin: 1rem auto;
#                   background: #2a2a2a;
#                   color: #fff;
#                   border: none;
#                   padding: 0.5rem;
#                   border-radius: 8px;
#                 }
#                 button {
#                   padding: 0.5rem 1rem;
#                   background: #4285F4;
#                   color: #fff;
#                   border: none;
#                   border-radius: 8px;
#                   cursor: pointer;
#                   font-size: 1rem;
#                 }
#                 button:hover {
#                   background: #357ae8;
#                 }
#             </style>
#         </head>
#         <body>
#             <div class="container">
#                 <h2>Import Judges from XLSX</h2>
#                 <form method="POST" enctype="multipart/form-data">
#                     <input type="file" name="file" accept=".xlsx" required>
#                     <button type="submit">Upload</button>
#                 </form>
#             </div>
#         </body>
#         </html>
#         ''')
    
#     # POST: Process the uploaded file
#     uploaded_file = request.files.get('file')
#     if not uploaded_file:
#         return "No file uploaded", 400

#     try:
#         wb = openpyxl.load_workbook(uploaded_file)
#         sheet = wb.active

#         # Read header row (assume first row is header)
#         headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
#         poster_cols = {}
#         poster_title_cols = {}
#         for i in range(1, 7):
#             try:
#                 poster_cols[i] = headers.index(f"poster-{i}")
#             except ValueError:
#                 poster_cols[i] = None
#             try:
#                 poster_title_cols[i] = headers.index(f"poster-{i}-title")
#             except ValueError:
#                 poster_title_cols[i] = None

#         try:
#             email_index = headers.index("Email")
#             fname_index = headers.index("Judge FirstName")
#             lname_index = headers.index("Judge LastName")
#         except ValueError as e:
#             return f"Missing required column in header: {str(e)}", 400

#         imported_count = 0
#         db = sqlite3.connect(DATABASE)
#         db.row_factory = sqlite3.Row
#         for row in sheet.iter_rows(min_row=2):
#             email = row[email_index].value
#             if not email:
#                 continue
#             first_name = row[fname_index].value or ""
#             last_name = row[lname_index].value or ""
#             full_name = f"{first_name.strip()} {last_name.strip()}".strip()

#             assigned_posters = []
#             poster_titles = {}
#             for i in range(1, 7):
#                 p_col = poster_cols.get(i)
#                 if p_col is not None:
#                     cell_value = row[p_col].value
#                     if is_valid_poster(cell_value):
#                         try:
#                             poster_num = int(float(cell_value))
#                         except Exception:
#                             continue
#                         poster_id = f"poster{poster_num}"
#                         if poster_id not in assigned_posters:
#                             assigned_posters.append(poster_id)
#                             t_col = poster_title_cols.get(i)
#                             title = row[t_col].value if t_col is not None else None
#                             if title:
#                                 poster_titles[poster_id] = title.strip()

#             if not assigned_posters:
#                 continue

#             db.execute("""
#                 INSERT INTO judges (email, name, assigned_posters, assigned_poster_titles)
#                 VALUES (?, ?, ?, ?)
#                 ON CONFLICT(email) DO UPDATE SET
#                     name = excluded.name,
#                     assigned_posters = excluded.assigned_posters,
#                     assigned_poster_titles = excluded.assigned_poster_titles
#             """, (email.strip(), full_name, json.dumps(assigned_posters), json.dumps(poster_titles)))
#             imported_count += 1

#         db.commit()
#         db.close()

#         # Return a dark-themed success page instead of plain text.
#         return render_template_string('''
#         <!DOCTYPE html>
#         <html lang="en">
#         <head>
#            <meta charset="UTF-8">
#            <meta name="viewport" content="width=device-width, initial-scale=1.0">
#            <title>Import Successful</title>
#            <link rel="stylesheet"
#                  href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;600&display=swap">
#            <style>
#               body {
#                  background-color: #121212;
#                  color: #fff;
#                  font-family: 'IBM Plex Sans', sans-serif;
#                  padding: 2rem;
#                  text-align: center;
#               }
#               .container {
#                  max-width: 500px;
#                  margin: auto;
#                  background: #1f1f1f;
#                  padding: 2rem;
#                  border-radius: 12px;
#                  box-shadow: 0 4px 6px rgba(0,0,0,0.3);
#               }
#               a.btn {
#                  display: inline-block;
#                  margin-top: 1rem;
#                  padding: 0.5rem 1rem;
#                  background: #4285F4;
#                  color: #fff;
#                  border-radius: 8px;
#                  text-decoration: none;
#                  font-size: 1rem;
#               }
#               a.btn:hover {
#                  background: #357ae8;
#               }
#            </style>
#         </head>
#         <body>
#            <div class="container">
#               <h2>Import Successful</h2>
#               <p>{{ imported_count }} judges imported successfully.</p>
#               <a href="{{ url_for('admin_dashboard') }}?key=adminsecret" class="btn">Back to Admin Dashboard</a>
#            </div>
#         </body>
#         </html>
#         ''', imported_count=imported_count), 200

#     except Exception as e:
#         return f"Error processing file: {str(e)}", 500

import openpyxl  # pip install openpyxl

def is_valid_poster(cell_value):
    """Return True if the cell value represents a valid poster number."""
    try:
        return cell_value is not None and (isinstance(cell_value, (int, float)) or str(cell_value).strip() != "")
    except Exception:
        return False

@app.route('/admin/import_judges', methods=['GET', 'POST'])
def import_judges():
    """
    Admin-only endpoint (accessible only if ?key=adminsecret is provided)
    that accepts an XLSX file upload containing judge and poster assignments.
    
    Expected header columns include:
      - Judge, Judge FirstName, Judge LastName, Email,
        poster-1, poster-1-title, poster-2, poster-2-title, ... poster-6, poster-6-title
      
    All judges are saved in the judges table.  
    Judges with no poster assignment will have an empty list in assigned_posters.
    """
    if request.args.get('key') != 'adminsecret':
        return "Unauthorized", 401

    if request.method == 'GET':
        return render_template_string('''
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Import Judges</title>
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <link rel="stylesheet"
                  href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;600&display=swap">
            <style>
                body {
                  background-color: #121212;
                  color: #fff;
                  font-family: 'IBM Plex Sans', sans-serif;
                  margin: 0;
                  padding: 2rem;
                }
                .container {
                  max-width: 500px;
                  margin: auto;
                  background: #1f1f1f;
                  padding: 2rem;
                  border-radius: 12px;
                  box-shadow: 0 4px 6px rgba(0,0,0,0.3);
                  text-align: center;
                }
                input[type="file"] {
                  display: block;
                  margin: 1rem auto;
                  background: #2a2a2a;
                  color: #fff;
                  border: none;
                  padding: 0.5rem;
                  border-radius: 8px;
                }
                button {
                  padding: 0.5rem 1rem;
                  background: #4285F4;
                  color: #fff;
                  border: none;
                  border-radius: 8px;
                  cursor: pointer;
                  font-size: 1rem;
                }
                button:hover {
                  background: #357ae8;
                }
            </style>
        </head>
        <body>
            <div class="container">
                <h2>Import Judges from XLSX</h2>
                <form method="POST" enctype="multipart/form-data">
                    <input type="file" name="file" accept=".xlsx" required>
                    <button type="submit">Upload</button>
                </form>
            </div>
        </body>
        </html>
        ''')
    
    # POST: Process the file
    uploaded_file = request.files.get('file')
    if not uploaded_file:
        return "No file uploaded", 400

    try:
        wb = openpyxl.load_workbook(uploaded_file)
        sheet = wb.active

        # Read header row
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        
        # Required columns: Judge, Judge FirstName, Judge LastName, Email
        try:
            judge_id_index = headers.index("Judge")
            fname_index = headers.index("Judge FirstName")
            lname_index = headers.index("Judge LastName")
            email_index = headers.index("Email")
        except ValueError as e:
            return f"Missing required column in header: {str(e)}", 400

        # Determine indexes for poster and poster-title columns (poster-1 .. poster-6)
        poster_cols = {}
        poster_title_cols = {}
        for i in range(1, 7):
            try:
                poster_cols[i] = headers.index(f"poster-{i}")
            except ValueError:
                poster_cols[i] = None
            try:
                poster_title_cols[i] = headers.index(f"poster-{i}-title")
            except ValueError:
                poster_title_cols[i] = None

        imported_count = 0
        db = sqlite3.connect(DATABASE)
        db.row_factory = sqlite3.Row
        for row in sheet.iter_rows(min_row=2):
            judge_id = row[judge_id_index].value
            if judge_id is None:
                continue
            first_name = row[fname_index].value or ""
            last_name = row[lname_index].value or ""
            full_name = f"{first_name.strip()} {last_name.strip()}".strip()
            email = row[email_index].value
            if not email:
                continue

            assigned_posters = []
            poster_titles = {}
            for i in range(1, 7):
                p_col = poster_cols.get(i)
                if p_col is not None:
                    cell_value = row[p_col].value
                    if is_valid_poster(cell_value):
                        try:
                            poster_num = int(float(cell_value))
                        except Exception:
                            continue
                        poster_id = f"poster{poster_num}"
                        if poster_id not in assigned_posters:
                            assigned_posters.append(poster_id)
                            t_col = poster_title_cols.get(i)
                            title = row[t_col].value if t_col is not None else None
                            if title:
                                poster_titles[poster_id] = title.strip()

            # Save judge in the judges table even if no posters are assigned.
            db.execute("""
                INSERT INTO judges (email, name, assigned_posters, assigned_poster_titles)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(email) DO UPDATE SET
                    name = excluded.name,
                    assigned_posters = excluded.assigned_posters,
                    assigned_poster_titles = excluded.assigned_poster_titles
            """, (email.strip(), full_name, json.dumps(assigned_posters), json.dumps(poster_titles)))
            # Save judge info with provided judge_id
            db.execute("""
                INSERT OR REPLACE INTO judge_info (judge_id, name, email)
                VALUES (?, ?, ?)
            """, (judge_id, full_name, email.strip()))
            imported_count += 1

        db.commit()
        db.close()

        return render_template_string('''
        <!DOCTYPE html>
        <html lang="en">
        <head>
           <meta charset="UTF-8">
           <meta name="viewport" content="width=device-width, initial-scale=1.0">
           <title>Import Successful</title>
           <link rel="stylesheet"
                 href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;600&display=swap">
           <style>
              body {
                 background-color: #121212;
                 color: #fff;
                 font-family: 'IBM Plex Sans', sans-serif;
                 padding: 2rem;
                 text-align: center;
              }
              .container {
                 max-width: 500px;
                 margin: auto;
                 background: #1f1f1f;
                 padding: 2rem;
                 border-radius: 12px;
                 box-shadow: 0 4px 6px rgba(0,0,0,0.3);
              }
              a.btn {
                 display: inline-block;
                 margin-top: 1rem;
                 padding: 0.5rem 1rem;
                 background: #4285F4;
                 color: #fff;
                 border-radius: 8px;
                 text-decoration: none;
                 font-size: 1rem;
              }
              a.btn:hover {
                 background: #357ae8;
              }
           </style>
        </head>
        <body>
           <div class="container">
              <h2>Import Successful</h2>
              <p>{{ imported_count }} judges imported successfully.</p>
              <a href="{{ url_for('admin_dashboard') }}?key=adminsecret" class="btn">Back to Admin Dashboard</a>
           </div>
        </body>
        </html>
        ''', imported_count=imported_count), 200

    except Exception as e:
        return f"Error processing file: {str(e)}", 500

@app.route('/admin/export_score_matrix')
def export_score_matrix():
    """
    Admin-only endpoint to export a score matrix in XLSX format.
    The matrix rows represent posters (e.g. "Poster-1", "Poster-2", …)
    and the columns represent judges (e.g. "Judge-1", "Judge-2", …).
    Each cell shows the score given by that judge to that poster (or 0 if not scored).
    Access is protected by a simple query parameter key (e.g., ?key=adminsecret).
    """
    # Check admin access via the key
    if request.args.get('key') != 'adminsecret':
        return "Unauthorized", 401

    db = get_db()

    # 1. Get the list of judges (using the judge_info table) sorted by judge_id.
    judges = db.execute("SELECT judge_id, email FROM judge_info ORDER BY judge_id").fetchall()
    judge_list = [{'id': judge['judge_id'], 'email': judge['email']} for judge in judges]

    # 2. Build a set of all poster IDs.
    #    We include poster ids from the scores table and from the judges' assigned_posters.
    poster_set = set()

    # From the scores table:
    score_rows = db.execute("SELECT DISTINCT poster_id FROM scores").fetchall()
    for row in score_rows:
        if row['poster_id']:
            poster_set.add(row['poster_id'])

    # From each judge's assigned_posters (stored as JSON):
    judges_all = db.execute("SELECT assigned_posters FROM judges").fetchall()
    for row in judges_all:
        assigned = row['assigned_posters']
        if assigned:
            try:
                posters = json.loads(assigned)
                for p in posters:
                    poster_set.add(p)
            except Exception:
                pass

    # Sort the poster list by the numeric portion.
    def poster_sort_key(p):
        try:
            return int(p.replace("poster", ""))
        except Exception:
            return 0
    poster_list = sorted(list(poster_set), key=poster_sort_key)

    # 3. Build a dictionary of scores keyed by (poster_id, judge_email).
    #    If a judge has not scored a poster, the default will be 0.
    score_rows = db.execute("SELECT judge_email, poster_id, score FROM scores").fetchall()
    score_dict = {}
    for row in score_rows:
        score_dict[(row['poster_id'], row['judge_email'])] = row['score']

    # 4. Create a new XLSX workbook and worksheet using openpyxl.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Score Matrix"

    # Write header row.
    # The first cell in the header row is left blank; then one column per judge.
    ws.cell(row=1, column=1, value="")  # Top-left cell is blank.
    col_index = 2
    for judge in judge_list:
        header = f"Judge-{judge['id']}"
        ws.cell(row=1, column=col_index, value=header)
        col_index += 1

    # 5. Write each poster row.
    # The first column contains the poster label (e.g., "Poster-1"),
    # and the subsequent columns contain the score for each judge (or 0 if no score exists).
    row_index = 2
    for poster in poster_list:
        try:
            poster_num = int(poster.replace("poster", ""))
        except Exception:
            poster_num = poster
        ws.cell(row=row_index, column=1, value=f"Poster-{poster_num}")
        col_index = 2
        for judge in judge_list:
            # Get the score if it exists; otherwise use 0.
            score_value = score_dict.get((poster, judge['email']), 0)
            ws.cell(row=row_index, column=col_index, value=score_value)
            col_index += 1
        row_index += 1

    # 6. Save the workbook to a BytesIO stream.
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 7. Return the XLSX file as an attachment.
    return send_file(output,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name="score_matrix.xlsx")


@app.route('/qr/<poster_id>')
def qr_redirect(poster_id):
    """
    QR endpoint: When a judge scans a QR code pointing to /qr/<poster_id>,
    if logged in, they are immediately redirected to /score/<poster_id>.
    Otherwise, the intended URL is saved in the session and the judge is sent to login.
    """
    if 'user' in session:
        return redirect(url_for('score', poster_id=poster_id))
    else:
        # Save the intended redirect URL for after login
        session['redirect_after_login'] = url_for('score', poster_id=poster_id)
        return redirect(url_for('login'))



# ===============================
# Admin Export Endpoint
# ===============================
@app.route('/export')
def export():
    """
    Exports the complete scores as a CSV.
    For demo purposes, access is protected by a simple key in the URL (e.g., /export?key=adminsecret).
    """
    key = request.args.get('key')
    if key != 'adminsecret':  # Replace with a secure method in production
        return "Unauthorized", 401
    db = get_db()
    cur = db.execute("SELECT * FROM scores")
    rows = cur.fetchall()
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["judge_email", "poster_id", "score"])
    for row in rows:
        writer.writerow([row['judge_email'], row['poster_id'], row['score']])
    output.seek(0)
    return send_file(BytesIO(output.getvalue().encode()),
                 mimetype='text/csv',
                 as_attachment=True,
                 download_name='scores.csv')
@app.route('/admin/view_scores', methods=['GET'])
def admin_view_scores():
    # Check admin access using query parameter
    if request.args.get('key') != 'adminsecret':
        return "Unauthorized", 401

    db = get_db()
    # Retrieve all scores ordered by judge_email and poster_id for clarity
    cur = db.execute("SELECT id, judge_email, poster_id, score FROM scores ORDER BY judge_email, poster_id")
    scores = cur.fetchall()

    return render_template_string('''
    <!DOCTYPE html>
    <html lang="en">
    <head>
      <meta charset="UTF-8">
      <title>View/Edit Scores</title>
      <meta name="viewport" content="width=device-width, initial-scale=1.0">
      <link rel="stylesheet"
            href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;600&display=swap">
      <style>
        body {
          background-color: #121212;
          color: #fff;
          font-family: 'IBM Plex Sans', sans-serif;
          padding: 2rem;
        }
        table {
          width: 100%;
          border-collapse: collapse;
          margin-bottom: 2rem;
        }
        th, td {
          border: 1px solid #2a2a2a;
          padding: 0.5rem;
          text-align: left;
        }
        th {
          background-color: #1f1f1f;
        }
        tr:nth-child(even) {
          background-color: #1f1f1f;
        }
        .btn {
          background-color: #4285F4;
          color: #fff;
          padding: 0.5rem 1rem;
          border: none;
          border-radius: 8px;
          text-decoration: none;
          cursor: pointer;
        }
        .btn:hover {
          background-color: #357ae8;
        }
        .form-inline {
          display: flex;
          gap: 0.5rem;
          align-items: center;
        }
        input[type="number"] {
          width: 80px;
          padding: 0.3rem;
          border-radius: 4px;
          border: none;
          text-align: center;
        }
      </style>
    </head>
    <body>
      <h1>View/Edit Scores</h1>
      <table>
        <thead>
          <tr>
            <th>ID</th>
            <th>Judge Email</th>
            <th>Poster ID</th>
            <th>Score</th>
            <th>Edit</th>
          </tr>
        </thead>
        <tbody>
          {% for score in scores %}
          <tr>
            <td>{{ score['id'] }}</td>
            <td>{{ score['judge_email'] }}</td>
            <td>{{ score['poster_id'] }}</td>
            <td>{{ score['score'] }}</td>
            <td>
              <form class="form-inline" method="POST" action="{{ url_for('admin_edit_score') }}?key=adminsecret">
                <input type="hidden" name="score_id" value="{{ score['id'] }}">
                <input type="number" name="new_score" min="0" max="10" placeholder="0-10" required>
                <button type="submit" class="btn">Update</button>
              </form>
            </td>
          </tr>
          {% endfor %}
        </tbody>
      </table>
      <a href="{{ url_for('admin_dashboard') }}?key=adminsecret" class="btn">Back to Admin Dashboard</a>
    </body>
    </html>
    ''', scores=scores)

@app.route('/admin/edit_score', methods=['POST'])
def admin_edit_score():
    if request.args.get('key') != 'adminsecret':
        return "Unauthorized", 401

    score_id = request.form.get('score_id')
    new_score = request.form.get('new_score')
    try:
        new_score = int(new_score)
        if new_score < 0 or new_score > 10:
            flash("Score must be between 0 and 10.")
            return redirect(url_for('admin_view_scores') + "?key=adminsecret")
    except ValueError:
        flash("Invalid score input.")
        return redirect(url_for('admin_view_scores') + "?key=adminsecret")

    db = get_db()
    # Update the score in the scores table
    db.execute("UPDATE scores SET score = ? WHERE id = ?", (new_score, score_id))
    db.commit()
    flash("Score updated.")
    return redirect(url_for('admin_view_scores') + "?key=adminsecret")

@app.route('/admin/reset_scores', methods=['GET', 'POST'])
def reset_scores():
    if request.args.get('key') != 'adminsecret':
        return "Unauthorized", 401

    if request.method == 'GET':
        return render_template_string('''
         <!DOCTYPE html>
         <html lang="en">
         <head>
             <meta charset="UTF-8">
             <title>Reset Scores</title>
             <meta name="viewport" content="width=device-width, initial-scale=1.0">
             <link rel="stylesheet"
                   href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;600&display=swap">
             <style>
                 body { background-color: #121212; color: #fff; font-family: 'IBM Plex Sans', sans-serif; padding: 2rem; }
                 .container { max-width: 500px; margin: auto; background: #1f1f1f; padding: 1rem; border-radius: 8px; text-align: center; }
                 .btn { background-color: #4285F4; color: #fff; padding: 0.5rem 1rem; border: none; border-radius: 8px; text-decoration: none; cursor: pointer; }
                 .btn:hover { background-color: #357ae8; }
             </style>
         </head>
         <body>
             <div class="container">
                 <h2>Reset Score Data</h2>
                 <p>Are you sure you want to reset all score data?<br>
                    This will remove all entries from the scores and score_changes tables.</p>
                 <form method="POST">
                     <button type="submit" class="btn">Reset Scores</button>
                 </form>
                 <br>
                 <a href="{{ url_for('admin_dashboard') }}?key=adminsecret" class="btn">Cancel</a>
             </div>
         </body>
         </html>
        ''')
    # POST: Reset the scores and score_changes tables
    db = get_db()
    db.execute("DELETE FROM scores")
    db.execute("DELETE FROM score_changes")
    db.commit()
    flash("All score data has been reset.")
    return redirect(url_for('admin_dashboard') + "?key=adminsecret")


@app.route('/admin/generate_all_qr', methods=['GET'])
def admin_generate_all_qr():
    if request.args.get('key') != 'adminsecret':
        return "Unauthorized", 401

    db = get_db()
    cur = db.execute("SELECT assigned_posters, assigned_poster_titles FROM judges")
    unique_posters = {}
    for row in cur.fetchall():
        # Decode the assigned posters (JSON list) and poster titles (JSON dict)
        if row['assigned_posters']:
            try:
                posters = json.loads(row['assigned_posters'])
            except Exception:
                posters = []
        else:
            posters = []
        if row['assigned_poster_titles']:
            try:
                titles = json.loads(row['assigned_poster_titles'])
            except Exception:
                titles = {}
        else:
            titles = {}
        for poster in posters:
            # Add the poster if not already included.
            if poster not in unique_posters:
                unique_posters[poster] = titles.get(poster, "No Title")

    return render_template_string('''
    <!DOCTYPE html>
    <html lang="en">
    <head>
      <meta charset="UTF-8">
      <title>QR Codes for All Posters</title>
      <meta name="viewport" content="width=device-width, initial-scale=1.0">
      <link rel="stylesheet" 
            href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;600&display=swap">
      <style>
         body {
            background-color: #121212;
            color: #fff;
            font-family: 'IBM Plex Sans', sans-serif;
            padding: 2rem;
         }
         .container {
            max-width: 900px;
            margin: auto;
            background: #1f1f1f;
            padding: 2rem;
            border-radius: 12px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.3);
         }
         .grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 1rem;
         }
         .poster-card {
            border: 1px solid #2a2a2a;
            padding: 1rem;
            text-align: center;
            border-radius: 8px;
            background-color: #121212;
         }
         .poster-card h3 {
            margin-bottom: 0.5rem;
         }
         .poster-card img {
            margin-top: 0.5rem;
            width: 150px;
            height: 150px;
         }
         .btn {
            display: inline-block;
            margin-top: 1rem;
            padding: 0.5rem 1rem;
            background: #4285F4;
            color: #fff;
            border-radius: 8px;
            text-decoration: none;
            font-size: 1rem;
            cursor: pointer;
         }
         .btn:hover {
            background: #357ae8;
         }
         .print-btn {
            background: #5cb85c;
         }
         .print-btn:hover {
            background: #4cae4c;
         }
      </style>
      <script>
         function printPage() {
           window.print();
         }
      </script>
    </head>
    <body>
      <div class="container">
         <h2>QR Codes for All Posters</h2>
         <div class="grid">
         {% for poster_id, title in unique_posters.items() %}
            <div class="poster-card">
               <h3>{{ poster_id[6:] }}: {{ title }}</h3>
               <img src="{{ url_for('generate_qr', poster_id=poster_id, _external=True) }}" alt="QR for {{ poster_id }}">
            </div>
         {% endfor %}
         </div>
         <button class="btn print-btn" onclick="printPage()">Print QR Codes</button>
         <br>
         <a href="{{ url_for('admin_dashboard') }}?key=adminsecret" class="btn">Back to Admin Dashboard</a>
      </div>
    </body>
    </html>
    ''', unique_posters=unique_posters)


# @app.route('/admin/dashboard')
# def admin_dashboard():
#     if request.args.get('key') != 'adminsecret':
#         return "Unauthorized", 401

#     return render_template_string('''
#     <!DOCTYPE html>
#     <html lang="en">
#     <head>
#       <meta charset="UTF-8">
#       <title>Admin Dashboard</title>
#       <meta name="viewport" content="width=device-width, initial-scale=1.0">
#       <link rel="stylesheet"
#             href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;600&display=swap">
#       <style>
#         body {
#           background-color: #121212;
#           color: #fff;
#           font-family: 'IBM Plex Sans', sans-serif;
#           margin: 0;
#           padding: 0;
#         }
#         .header {
#           background-color: #1f1f1f;
#           padding: 2rem 1.5rem;
#           text-align: center;
#           border-bottom-left-radius: 24px;
#           border-bottom-right-radius: 24px;
#         }
#         .header h1 {
#           font-size: 2rem;
#           margin-bottom: 0.5rem;
#         }
#         .container {
#           padding: 2rem;
#           max-width: 800px;
#           margin: 2rem auto;
#         }
#         .dashboard-menu {
#           display: flex;
#           flex-wrap: wrap;
#           gap: 1rem;
#           justify-content: center;
#         }
#         .menu-item {
#           background-color: #1f1f1f;
#           border-radius: 12px;
#           padding: 1rem;
#           width: 250px;
#           text-align: center;
#           box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
#           transition: background-color 0.2s ease;
#         }
#         .menu-item:hover {
#           background-color: #2a2a2a;
#         }
#         .menu-item h2 {
#           font-size: 1.5rem;
#           margin-bottom: 0.5rem;
#         }
#         .menu-item p {
#           font-size: 1rem;
#           margin-bottom: 1rem;
#         }
#         .menu-item a {
#           display: inline-block;
#           margin-top: 0.5rem;
#           color: #4285F4;
#           text-decoration: none;
#           font-weight: 600;
#           padding: 0.5rem 1rem;
#           border: 2px solid #4285F4;
#           border-radius: 8px;
#           transition: background-color 0.2s ease;
#         }
#         .menu-item a:hover {
#           background-color: #4285F4;
#           color: #fff;
#         }
#       </style>
#     </head>
#     <body>
#       <div class="header">
#         <h1>Admin Dashboard</h1>
#         <p>Manage all admin operations from here</p>
#       </div>
#       <div class="container">
#         <div class="dashboard-menu">
#           <div class="menu-item">
#             <h2>Import Judges</h2>
#             <p>Upload an XLSX file to import judges and poster assignments.</p>
#             <a href="{{ url_for('import_judges') }}?key=adminsecret">Go to Import</a>
#           </div>
#           <div class="menu-item">
#             <h2>Remove Judges</h2>
#             <p>Remove all current judges from the system.</p>
#             <a href="{{ url_for('remove_judges') }}?key=adminsecret">Remove Judges</a>
#           </div>
#           <div class="menu-item">
#             <h2>Export Scores</h2>
#             <p>Download all scores as a CSV file.</p>
#             <a href="{{ url_for('export') }}?key=adminsecret">Export CSV</a>
#           </div>
#           <div class="menu-item">
#             <h2>View/Edit Scores</h2>
#             <p>Review and update scores submitted by judges.</p>
#             <a href="{{ url_for('admin_view_scores') }}?key=adminsecret">View/Edit Scores</a>
#           </div>
#           <div class="menu-item">
#             <h2>Reset Scores</h2>
#             <p>Clear all score data and logs for a new round.</p>
#             <a href="{{ url_for('reset_scores') }}?key=adminsecret">Reset Scores</a>
#           </div>
#           <div class="menu-item">
#             <h2>Generate All QR Codes</h2>
#             <p>Generate QR codes for all unique poster assignments.</p>
#             <a href="{{ url_for('admin_generate_all_qr') }}?key=adminsecret">Generate All QR Codes</a>
#           </div>
#         </div>
#       </div>
#     </body>
#     </html>
#     ''')

@app.route('/admin/dashboard')
def admin_dashboard():
    if request.args.get('key') != 'adminsecret':
        return "Unauthorized", 401

    return render_template_string('''
    <!DOCTYPE html>
    <html lang="en">
    <head>
      <meta charset="UTF-8">
      <title>Admin Dashboard</title>
      <meta name="viewport" content="width=device-width, initial-scale=1.0">
      <link rel="stylesheet"
            href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;600&display=swap">
      <style>
        body {
          background-color: #121212;
          color: #fff;
          font-family: 'IBM Plex Sans', sans-serif;
          margin: 0;
          padding: 0;
        }
        .header {
          background-color: #1f1f1f;
          padding: 2rem 1.5rem;
          text-align: center;
          border-bottom-left-radius: 24px;
          border-bottom-right-radius: 24px;
        }
        .header h1 {
          font-size: 2rem;
          margin-bottom: 0.5rem;
        }
        .container {
          padding: 2rem;
          max-width: 800px;
          margin: 2rem auto;
        }
        .dashboard-menu {
          display: flex;
          flex-wrap: wrap;
          gap: 1rem;
          justify-content: center;
        }
        .menu-item {
          background-color: #1f1f1f;
          border-radius: 12px;
          padding: 1rem;
          width: 250px;
          text-align: center;
          box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
          transition: background-color 0.2s ease;
        }
        .menu-item:hover {
          background-color: #2a2a2a;
        }
        .menu-item h2 {
          font-size: 1.5rem;
          margin-bottom: 0.5rem;
        }
        .menu-item p {
          font-size: 1rem;
          margin-bottom: 1rem;
        }
        .menu-item a {
          display: inline-block;
          margin-top: 0.5rem;
          color: #4285F4;
          text-decoration: none;
          font-weight: 600;
          padding: 0.5rem 1rem;
          border: 2px solid #4285F4;
          border-radius: 8px;
          transition: background-color 0.2s ease;
        }
        .menu-item a:hover {
          background-color: #4285F4;
          color: #fff;
        }
      </style>
    </head>
    <body>
      <div class="header">
        <h1>Admin Dashboard</h1>
        <p>Manage all admin operations from here</p>
      </div>
      <div class="container">
        <div class="dashboard-menu">
          <div class="menu-item">
            <h2>Import Judges</h2>
            <p>Upload an XLSX file to import judges and poster assignments.</p>
            <a href="{{ url_for('import_judges') }}?key=adminsecret">Go to Import</a>
          </div>
          <div class="menu-item">
            <h2>Remove Judges</h2>
            <p>Remove all current judges from the system.</p>
            <a href="{{ url_for('remove_judges') }}?key=adminsecret">Remove Judges</a>
          </div>
          <div class="menu-item">
            <h2>Export Scores</h2>
            <p>Download all scores as a CSV file.</p>
            <a href="{{ url_for('export') }}?key=adminsecret">Export CSV</a>
          </div>
          <div class="menu-item">
            <h2>View/Edit Scores</h2>
            <p>Review and update scores submitted by judges.</p>
            <a href="{{ url_for('admin_view_scores') }}?key=adminsecret">View/Edit Scores</a>
          </div>
          <div class="menu-item">
            <h2>Reset Scores</h2>
            <p>Clear all score data and logs for a new round.</p>
            <a href="{{ url_for('reset_scores') }}?key=adminsecret">Reset Scores</a>
          </div>
          <div class="menu-item">
            <h2>Generate All QR Codes</h2>
            <p>Generate QR codes for all unique poster assignments.</p>
            <a href="{{ url_for('admin_generate_all_qr') }}?key=adminsecret">Generate All QR Codes</a>
          </div>
          <div class="menu-item">
            <h2>Export Score Matrix</h2>
            <p>Download the score matrix as an XLSX file.</p>
            <a href="{{ url_for('export_score_matrix') }}?key=adminsecret">Export Matrix</a>
          </div>
        </div>
      </div>
    </body>
    </html>
    ''')




@app.route('/generate_qr/<poster_id>')
def generate_qr(poster_id):
    """
    Generates a QR code image for a given poster_id.
    The QR code encodes the URL for the /qr/<poster_id> endpoint.
    """
    # Construct the target URL that judges will be redirected to (i.e., /qr/<poster_id>)
    target_url = url_for('qr_redirect', poster_id=poster_id, _external=True)
    
    # Generate the QR code using the qrcode library
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(target_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Save image to a BytesIO object and send it as a PNG image
    img_io = BytesIO()
    img.save(img_io, 'PNG')
    img_io.seek(0)
    return send_file(img_io, mimetype='image/png')




# ===============================
# Run the Application
# ===============================
if __name__ == '__main__':
    app.run(debug=True)
